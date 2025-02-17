import openpyxl
import json
wb= openpyxl.load_workbook(r"/Users/hemav/Downloads/file.xlsx")
sheet= wb.active
data={}
columns_no= sheet.max_column
for i in range(1,columns_no):
    menu={}
    breakfast=[]
    lunch=[]
    dinner=[]
    def meal_section_end(start_row, column_index):
        for row in range(start_row, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=column_index).value
            if cell_value is not None and str(cell_value).strip().upper() in ["SATURDAY", "SUNDAY", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]:
                return row - 1
        return sheet.max_row
    breakfast_end = mealsection_end(4, i)
    lunch_start = breakfast_end + 2
    lunch_end = meal_section_end(lunch_start, i)
    dinner_start = lunch_end + 2
    dinner_end = meal_section_end(dinner_start, i)
    for j in range(4,breakfast_end+1):
        item=sheet.cell(row=j, column=i).value
        if "*" not in str(item) and item is not None:
            item_=" ".join(item.split())
            breakfast.append(item_)
    for j in range(lunch_start,lunch_end+1):
        item=sheet.cell(row=j, column=i).value
        if "*" not in str(item) and item is not None:
            item_=" ".join(item.split())
            lunch.append(item_)
    for j in range(dinner_start,dinner_end+1):
        item=sheet.cell(row=j, column=i).value
        if "*" not in str(item) and item is not None:
            item_=" ".join(item.split())
            dinner.append(item_)
    menu["BREAKFAST"]=breakfast
    menu["LUNCH"]=lunch
    menu["DINNER"]=dinner
    if i<10:
        data["2025-02-0"+str(i)]=menu
    else:
        data["2025-02-"+str(i)]=menu
f=open('//hemav//Downloads//mess_menu.json',"w")
json.dump(data,f,indent=2)

    

        
        
    
    


    

        
        
    
    
